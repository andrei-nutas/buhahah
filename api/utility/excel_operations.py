import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

from api.utility.excel_attributes import ExcelAttributes

def update_excel_with_extracted_content(excel_output_path, json_data):
    workbook = get_or_create_workbook(excel_output_path)
    worksheet = get_or_create_sheet(workbook, json_data.get("client_name"))
    remove_default_sheet(workbook)

    append_data_to_sheet(worksheet, json_data)

    workbook.save(excel_output_path)
    print(f"Excel file updated at {excel_output_path}")

def get_or_create_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)
    else:
        return Workbook()

def get_or_create_sheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(title = sheet_name)
        initialize_first_column_with_attributes_names(worksheet)
        return worksheet

def initialize_first_column_with_attributes_names(worksheet):
    second_row = 2
    for row_number, attribute_name in enumerate(ExcelAttributes, second_row):
        worksheet.cell(row = row_number, column = 1, value = attribute_name.value)

def remove_default_sheet(workbook):
    default_sheet_name = "Sheet"
    if default_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[default_sheet_name])

def append_data_to_sheet(worksheet, data):
    next_empty_column = worksheet.max_column + 1
    add_timestamp_as_header(next_empty_column, worksheet)

    second_row = 2
    for row_number, attribute_name in enumerate(ExcelAttributes, start = second_row):
        value = data.get(attribute_name.value)
        if value is not None:
            worksheet.cell(row = row_number, column = next_empty_column, value = value)

def add_timestamp_as_header(next_empty_column, worksheet):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    worksheet.cell(row = 1, column = next_empty_column, value = timestamp)
